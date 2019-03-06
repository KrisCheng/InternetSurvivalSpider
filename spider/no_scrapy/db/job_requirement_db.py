#!/usr/bin/python
# -*- coding: utf-8 -*-
# Author: kris_peng
# Created on 2019/2/27

import os
from datetime import datetime
import dateutil.parser
from openpyxl import load_workbook

from model.job_requirement import JobRequirement
from sqlalchemy import Column, String, Integer, DateTime, create_engine
from sqlalchemy.orm import sessionmaker
from config.config import *
from util.util import *

def main_task():

    engine = create_engine(MYSQL_DATABASE_URI)
    DBSession = sessionmaker(bind=engine)
    session = DBSession()

    list = []

    file_list = file_list_end_with(file_end="job_details.xlsx")
    for file in file_list:
        print("载入 " + file)
        wb = load_workbook(filename=os.path.join(DATA_BASE_PATH, file), data_only=True, read_only=True)
        ws = wb.worksheets[0]

        idx = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(row_offset=1):
            dict = {}
            for cell in row:
                cell_value = cell.value
                if cell_value is not None:
                    cell_key = idx[cell.column - 1]
                    if isinstance(cell_value, datetime):
                        cell_value = cell_value.strftime("%Y-%m-%d")
                    elif "日期" in cell_key:
                        cell_value = dateutil.parser.parse(cell_value).date().strftime("%Y-%m-%d")
                    dict[cell_key] = str(cell_value)
            list.append(dict)

    all_job_relation = fetch_all_jobrequirement(session=session)
    exist_jobcode_list = []
    for job_code in all_job_relation:
        exist_jobcode_list.append(job_code.job_code)
    current_jobcode_list = []

    for job_item in list:
        try:
            if job_item["职位编码"] in exist_jobcode_list or job_item["职位编码"] in current_jobcode_list:
                continue
            data_result = JobRequirement(
                job_code = job_item["职位编码"],
                title = job_item["职位类型"],
                property = job_item["工作性质"],
                experience = job_item["工作经验"],
                education = job_item["教育程度"],
                description = job_item["详情描述"],
                data_source = "Lagou")
            session.add(data_result)
            current_jobcode_list.append(job_item["职位编码"])

        except:
            print("Lagou Job Requirement Import ERROR.")
        finally:
            session.commit()
    session.close()
